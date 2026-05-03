import io
from datetime import datetime


class ExcelExporter:

    @staticmethod
    def export_conversations(queryset) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Conversations'

        # Header style
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='2E5FA3')

        headers = [
            'ID', 'Client', 'Email', 'Phone', 'Source',
            'Agent', 'Status', 'Platform', 'Created At', 'Updated At'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font    = header_font
            cell.fill    = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row, conv in enumerate(queryset, 2):
            ws.cell(row=row, column=1,  value=str(conv.id)[:8] + '...')
            ws.cell(row=row, column=2,  value=conv.client.full_name())
            ws.cell(row=row, column=3,  value=conv.client.email or '-')
            ws.cell(row=row, column=4,  value=conv.client.phone or '-')
            ws.cell(row=row, column=5,  value=conv.client.source)
            ws.cell(row=row, column=6,  value=conv.agent.full_name() if conv.agent else 'Unassigned')
            ws.cell(row=row, column=7,  value=conv.status)
            ws.cell(row=row, column=8,  value=conv.channel.platform if conv.channel else '-')
            ws.cell(row=row, column=9,  value=conv.created_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=10, value=conv.updated_at.strftime('%Y-%m-%d %H:%M'))

        # Auto-width columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def export_agents(agent_stats: list) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Agent Performance'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='2E5FA3')

        headers = [
            'Agent Name', 'Email', 'Total Handled',
            'Resolved', 'Open', 'Resolution Rate %', 'Current Status'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row, agent in enumerate(agent_stats, 2):
            ws.cell(row=row, column=1, value=agent['agent_name'])
            ws.cell(row=row, column=2, value=agent['agent_email'])
            ws.cell(row=row, column=3, value=agent['total_handled'])
            ws.cell(row=row, column=4, value=agent['resolved'])
            ws.cell(row=row, column=5, value=agent['open'])
            ws.cell(row=row, column=6, value=f"{agent['resolution_rate']}%")
            ws.cell(row=row, column=7, value=agent['current_status'])

        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def export_clients(queryset) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Clients'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='2E5FA3')

        headers = ['First Name', 'Last Name', 'Email', 'Phone', 'Source', 'Joined At']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row, client in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=client.first_name or '-')
            ws.cell(row=row, column=2, value=client.last_name  or '-')
            ws.cell(row=row, column=3, value=client.email      or '-')
            ws.cell(row=row, column=4, value=client.phone      or '-')
            ws.cell(row=row, column=5, value=client.source)
            ws.cell(row=row, column=6, value=client.created_at.strftime('%Y-%m-%d %H:%M'))

        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def export_full_report(overview, trends, agent_stats, platform_breakdown) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()

        # Sheet 1 — Overview
        ws1 = wb.active
        ws1.title = 'Overview'

        ws1['A1'] = 'PLATFORM OVERVIEW REPORT'
        ws1['A1'].font = Font(bold=True, size=14, color='2E5FA3')
        ws1['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'

        ws1['A4'] = 'CONVERSATIONS'
        ws1['A4'].font = Font(bold=True)
        ws1['A5']  = 'Total';        ws1['B5']  = overview['conversations']['total']
        ws1['A6']  = 'Open';         ws1['B6']  = overview['conversations']['open']
        ws1['A7']  = 'Pending';      ws1['B7']  = overview['conversations']['pending']
        ws1['A8']  = 'Resolved';     ws1['B8']  = overview['conversations']['resolved']

        ws1['A10'] = 'CLIENTS'
        ws1['A10'].font = Font(bold=True)
        ws1['A11'] = 'Total';        ws1['B11'] = overview['clients']['total']
        ws1['A12'] = 'New Today';    ws1['B12'] = overview['clients']['new_today']

        ws1['A14'] = 'AGENTS'
        ws1['A14'].font = Font(bold=True)
        ws1['A15'] = 'Total';        ws1['B15'] = overview['agents']['total']
        ws1['A16'] = 'Online';       ws1['B16'] = overview['agents']['online']
        ws1['A17'] = 'Busy';         ws1['B17'] = overview['agents']['busy']
        ws1['A18'] = 'Offline';      ws1['B18'] = overview['agents']['offline']

        # Sheet 2 — Agent Performance
        ws2 = wb.create_sheet('Agent Performance')
        headers = ['Agent', 'Email', 'Total', 'Resolved', 'Open', 'Resolution %', 'Status']
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(fill_type='solid', fgColor='2E5FA3')

        for row, agent in enumerate(agent_stats, 2):
            ws2.cell(row=row, column=1, value=agent['agent_name'])
            ws2.cell(row=row, column=2, value=agent['agent_email'])
            ws2.cell(row=row, column=3, value=agent['total_handled'])
            ws2.cell(row=row, column=4, value=agent['resolved'])
            ws2.cell(row=row, column=5, value=agent['open'])
            ws2.cell(row=row, column=6, value=f"{agent['resolution_rate']}%")
            ws2.cell(row=row, column=7, value=agent['current_status'])

        # Sheet 3 — Platform Breakdown
        ws3 = wb.create_sheet('By Platform')
        headers = ['Platform', 'Total', 'Open', 'Resolved']
        for col, h in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(fill_type='solid', fgColor='2E5FA3')

        for row, p in enumerate(platform_breakdown, 2):
            ws3.cell(row=row, column=1, value=p.get('channel__platform', '-'))
            ws3.cell(row=row, column=2, value=p.get('total', 0))
            ws3.cell(row=row, column=3, value=p.get('open', 0))
            ws3.cell(row=row, column=4, value=p.get('resolved', 0))

        # Sheet 4 — Trends
        ws4 = wb.create_sheet('Daily Trends')
        headers = ['Date', 'Total', 'Open', 'Resolved', 'Pending']
        for col, h in enumerate(headers, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(fill_type='solid', fgColor='2E5FA3')

        for row, trend in enumerate(trends, 2):
            ws4.cell(row=row, column=1, value=str(trend.get('period', '')))
            ws4.cell(row=row, column=2, value=trend.get('total', 0))
            ws4.cell(row=row, column=3, value=trend.get('open', 0))
            ws4.cell(row=row, column=4, value=trend.get('resolved', 0))
            ws4.cell(row=row, column=5, value=trend.get('pending', 0))

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


class PDFExporter:

    @staticmethod
    def export_conversations(queryset, title='Conversations Report') -> bytes:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm

        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        story  = []

        # Title
        story.append(Paragraph(title, styles['Title']))
        story.append(Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
        story.append(Spacer(1, 0.5 * cm))

        # Table data
        data = [['Client', 'Source', 'Agent', 'Status', 'Platform', 'Created']]

        for conv in queryset:
            data.append([
                conv.client.full_name()[:20],
                conv.client.source,
                conv.agent.full_name()[:15] if conv.agent else 'Unassigned',
                conv.status,
                conv.channel.platform if conv.channel else '-',
                conv.created_at.strftime('%Y-%m-%d'),
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E5FA3')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 10),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F9FC')]),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('FONTSIZE',   (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))

        story.append(table)
        doc.build(story)
        return buffer.getvalue()

    @staticmethod
    def export_agents(agent_stats: list, title='Agent Performance Report') -> bytes:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm

        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story  = []

        story.append(Paragraph(title, styles['Title']))
        story.append(Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
        story.append(Spacer(1, 0.5 * cm))

        data = [['Agent', 'Total', 'Resolved', 'Open', 'Resolution %', 'Status']]

        for agent in agent_stats:
            data.append([
                agent['agent_name'],
                str(agent['total_handled']),
                str(agent['resolved']),
                str(agent['open']),
                f"{agent['resolution_rate']}%",
                agent['current_status'],
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E5FA3')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F9FC')]),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('FONTSIZE',   (0, 0), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))

        story.append(table)
        doc.build(story)
        return buffer.getvalue()

    @staticmethod
    def export_full_report(overview, agent_stats, platform_breakdown) -> bytes:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm

        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story  = []

        story.append(Paragraph('Platform Analytics Report', styles['Title']))
        story.append(Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
        story.append(Spacer(1, 0.5 * cm))

        # Overview section
        story.append(Paragraph('Overview', styles['Heading1']))
        overview_data = [
            ['Metric', 'Value'],
            ['Total Conversations', overview['conversations']['total']],
            ['Open',                overview['conversations']['open']],
            ['Pending',             overview['conversations']['pending']],
            ['Resolved',            overview['conversations']['resolved']],
            ['Total Clients',       overview['clients']['total']],
            ['New Clients Today',   overview['clients']['new_today']],
            ['Total Agents',        overview['agents']['total']],
            ['Online Agents',       overview['agents']['online']],
            ['Busy Agents',         overview['agents']['busy']],
        ]
        ov_table = Table(overview_data, colWidths=[10*cm, 5*cm])
        ov_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E5FA3')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F9FC')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(ov_table)
        story.append(Spacer(1, 0.5 * cm))

        # Agent performance section
        story.append(Paragraph('Agent Performance', styles['Heading1']))
        agent_data = [['Agent', 'Total', 'Resolved', 'Resolution %', 'Status']]
        for agent in agent_stats:
            agent_data.append([
                agent['agent_name'],
                agent['total_handled'],
                agent['resolved'],
                f"{agent['resolution_rate']}%",
                agent['current_status'],
            ])
        ag_table = Table(agent_data)
        ag_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E5FA3')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F9FC')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(ag_table)

        doc.build(story)
        return buffer.getvalue()